from flask import Flask, request, render_template, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Nomes dos arquivos de planilha
FILE_NAME_ATENDIMENTO = 'atendimento_respostas.xlsx'
FILE_NAME_CLIENTE = 'cliente_respostas.xlsx'
FILE_NAME_CORRETOR = 'corretor_respostas.xlsx'
FILE_NAME_PONTO = 'ponto_respostas.xlsx'

def init_excel_file(file_name, headers):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_name)

def check_duplicate(file_name, cpf):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == cpf:
                return True
    return False

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/selecionar_vendedor', methods=['POST'])
def selecionar_vendedor():
    vendedor = request.form.get('vendedor')
    if vendedor not in ['vendedor1', 'vendedor2']:
        abort(404)  # Retorna erro 404 se o vendedor não for válido
    # Aqui você pode armazenar a seleção do vendedor na sessão ou no banco de dados se necessário
    # Exemplo: session['vendedor'] = vendedor
    return redirect(url_for('menu', vendedor=vendedor))

@app.route('/menu')
def menu():
    return render_template('menu.html')

@app.route('/atendimento')
def atendimento():
    return render_template('atendimento.html')

@app.route('/submit_atendimento_form', methods=['POST'])
def submit_atendimento_form():
    atendido = request.form['atendido']
    if atendido == "Sim":
        # Lógica para quando o cliente será atendido
        return redirect(url_for('cliente_form'))
    elif atendido == "Não":
        # Lógica para quando o cliente não será atendido
        return redirect(url_for('menu'))
    else:
        # Lógica para quando o cliente será um retorno
        return redirect(url_for('retorno_cliente'))

@app.route('/retorno_cliente')
def retorno_cliente():
    return render_template('retorno_cliente.html')

@app.route('/submit_retorno_form', methods=['POST'])
def submit_retorno_form():
    nome = request.form['nome']
    cpf = request.form['cpf']
    # Lógica para salvar os dados do cliente retornado
    return redirect(url_for('menu'))

@app.route('/cliente_form')
def cliente_form():
    return render_template('cliente_form.html')

@app.route('/submit_cliente_form', methods=['POST'])
def submit_cliente_form():
    nome = request.form['nome']
    telefone = request.form['telefone']
    email = request.form['email']
    data_visita = request.form['data_visita']
    corretor = request.form['corretor']
    imobiliaria = request.form['imobiliaria']
    como_soube = request.form['como_soube']

    # Aqui você pode salvar os dados no banco de dados
    cliente = Cliente(nome=nome, telefone=telefone, email=email, data_visita=data_visita, corretor=corretor, imobiliaria=imobiliaria, como_soube=como_soube)
    db.session.add(cliente)
    db.session.commit()

    return redirect(url_for('menu'))

@app.route('/corretor_form')
def corretor_form():
    return render_template('corretor_form.html')

@app.route('/submit_corretor_form', methods=['POST'])
def submit_corretor_form():
    cpf = request.form['cpf']
    nome = request.form['nome']
    imobiliaria = request.form['imobiliaria']
    
    if check_duplicate(FILE_NAME_CORRETOR, cpf):
        return "Erro: CPF já cadastrado."
    
    if not os.path.exists(FILE_NAME_CORRETOR):
        init_excel_file(FILE_NAME_CORRETOR, ['CPF', 'Nome', 'Imobiliária'])
    
    wb = load_workbook(FILE_NAME_CORRETOR)
    ws = wb.active
    ws.append([cpf, nome, imobiliaria])
    wb.save(FILE_NAME_CORRETOR)
    
    return redirect(url_for('menu'))

@app.route('/mesa')
def mesa():
    return render_template('mesa.html')

@app.route('/ponto_form')
def ponto_form():
    periodo = request.args.get('periodo')
    mesa = request.args.get('mesa')
    return render_template('ponto_form.html', periodo=periodo, mesa=mesa)

@app.route('/submit_ponto_form', methods=['POST'])
def submit_ponto_form():
    periodo = request.form['periodo']
    mesa = request.form['mesa']
    cpf = request.form['cpf']
    data = request.form['data']
    horario = request.form['horario']
    observacoes = request.form.get('observacoes', '')

    # Aqui você pode salvar esses dados no banco de dados
    # Exemplo:
    # novo_ponto = Ponto(periodo=periodo, mesa=mesa, cpf=cpf, data=data, horario=horario, observacoes=observacoes)
    # db.session.add(novo_ponto)
    # db.session.commit()

    return redirect(url_for('menu'))

if __name__ == '__main__':
    init_excel_file(FILE_NAME_ATENDIMENTO, ['Atendido', 'Data'])
    init_excel_file(FILE_NAME_CLIENTE, ['Nome', 'Telefone', 'E-mail', 'Data da Visita', 'Como Soube'])
    init_excel_file(FILE_NAME_CORRETOR, ['CPF', 'Nome', 'Imobiliária'])
    init_excel_file(FILE_NAME_PONTO, ['CPF', 'Data', 'Horário de Entrada', 'Horário de Saída'])
    app.run(debug=True)
